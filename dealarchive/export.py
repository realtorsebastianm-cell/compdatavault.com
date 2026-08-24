"""Export a broker's selected comps to an .xlsx workbook -- the "hand this
to a lender/client/appraiser" feature. Deliberately not a PDF: a comp set
almost always gets pasted into someone else's spreadsheet or CA next, and
.xlsx is the format that doesn't lose that round-trip.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from dealarchive.models import LeaseComp, SaleComp

SALE_COLUMNS = [
    ("Address", lambda c: c.address),
    ("City", lambda c: c.city),
    ("State", lambda c: c.state),
    ("Submarket", lambda c: c.submarket),
    ("Property Type", lambda c: c.property_type.value if c.property_type else None),
    ("Building SF", lambda c: float(c.building_sf) if c.building_sf is not None else None),
    ("Lot SF", lambda c: float(c.lot_sf) if c.lot_sf is not None else None),
    ("Price", lambda c: float(c.price) if c.price is not None else None),
    ("Price / SF", lambda c: float(c.price_per_sf) if c.price_per_sf is not None else None),
    ("Cap Rate", lambda c: float(c.cap_rate) if c.cap_rate is not None else None),
    ("Units", lambda c: float(c.num_units) if c.num_units is not None else None),
    ("Price / Unit", lambda c: float(c.price_per_unit) if c.price_per_unit is not None else None),
    ("Zoning", lambda c: c.zoning),
    ("Broker", lambda c: c.broker_name),
    ("Brokerage", lambda c: c.brokerage),
    ("Date Received", lambda c: c.date_received),
    ("Notes", lambda c: c.notes),
]

LEASE_COLUMNS = [
    ("Address", lambda c: c.address),
    ("City", lambda c: c.city),
    ("State", lambda c: c.state),
    ("Submarket", lambda c: c.submarket),
    ("Property Type", lambda c: c.property_type.value if c.property_type else None),
    ("Building SF", lambda c: float(c.building_sf) if c.building_sf is not None else None),
    ("Lot SF", lambda c: float(c.lot_sf) if c.lot_sf is not None else None),
    ("Rate", lambda c: float(c.rate) if c.rate is not None else None),
    ("Rate Type", lambda c: c.rate_type.value if c.rate_type else None),
    ("Term (months)", lambda c: c.term_months),
    ("Expenses", lambda c: c.expense_type.value if c.expense_type else None),
    ("Zoning", lambda c: c.zoning),
    ("Broker", lambda c: c.broker_name),
    ("Brokerage", lambda c: c.brokerage),
    ("Date Received", lambda c: c.date_received),
    ("Notes", lambda c: c.notes),
]


def _write_sheet(ws, columns, comps) -> None:
    header_font = Font(bold=True)
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

    for row_idx, comp in enumerate(comps, start=2):
        for col_idx, (_, getter) in enumerate(columns, start=1):
            value = getter(comp)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, str) and len(value) > 40:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, (label, _) in enumerate(columns, start=1):
        # Notes tends to run long; cap the width so it doesn't blow out the
        # sheet, everything else sizes to its header.
        width = 40 if label == "Notes" else max(12, len(label) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


def build_workbook(sale_comps: list[SaleComp], lease_comps: list[LeaseComp]) -> bytes:
    wb = Workbook()

    first_sheet_used = False
    if sale_comps:
        ws = wb.active
        ws.title = "Sale Comps"
        _write_sheet(ws, SALE_COLUMNS, sale_comps)
        first_sheet_used = True

    if lease_comps:
        ws = wb.active if not first_sheet_used else wb.create_sheet()
        ws.title = "Lease Comps"
        _write_sheet(ws, LEASE_COLUMNS, lease_comps)
        first_sheet_used = True

    if not first_sheet_used:
        wb.active.title = "Comps"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
